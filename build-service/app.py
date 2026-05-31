"""
app.py  -  PCAOB Rotation Build Service  v3
Kreit & Chiu CPA LLP

Review Comments Implementation (all 14 sections):
  S1/S3  - Rotation Dashboard filtered to active Bizinta clients only;
            EP column split into "EP as per Form AP" and "EP as per Bizinta"
  S2.2   - New tab: PCAOB Filings with Bizinta Status
  S2.3   - New tab: Active Bizinta Clients with No PCAOB Filing
  S4/S12 - All filtering now by Firm ID 6651 (done in index.js);
            Raw PCAOB Filings tab shows all records regardless of firm name variant
  S5     - Partner Summary rebuilt from corrected active-client scope
  S6     - Client Reconciliation logic corrected; Category 2 now populated correctly
  S7     - Name Changes tab restructured into 3 categories
  S8     - Legend & Notes updated: cooling-off note, dedup key examples, scope note
  S13    - Raw Bizinta tab flags active clients with no PCAOB mapping
  S14    - HTML dashboard reflects all above changes

Scope assumption (captured in banner and Legend tab):
  The Rotation Dashboard includes only clients that are BOTH Active in Bizinta
  AND have at least one Form AP filing on record with PCAOB (Firm ID 6651).
  Clients active in Bizinta with no PCAOB filing are listed in the
  'Active Bizinta - No Filing' tab for manual review.
  Clients with PCAOB filings but no longer active in Bizinta are excluded
  from the rotation dashboard as the partner rotation chain is considered broken.
"""

import os, io, csv, json, base64, urllib.request, urllib.error
from datetime import date, datetime as _dt
from collections import defaultdict
from functools import wraps
from flask import Flask, jsonify, request, abort
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

BIZINTA_SUBDOMAIN  = os.environ.get("BIZINTA_SUBDOMAIN", "")
BIZINTA_TOKEN      = os.environ.get("BIZINTA_TOKEN", "")
FILTER_SERVICE_URL = os.environ.get("FILTER_SERVICE_URL",
                     "https://pcaob-filter-service-production.up.railway.app/filter")
BUILD_API_KEY      = os.environ.get("BUILD_API_KEY", "")

def require_api_key(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if BUILD_API_KEY and request.headers.get("X-Api-Key") != BUILD_API_KEY:
            abort(401, "Invalid or missing API key")
        return f(*args, **kwargs)
    return decorated

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
F         = "Arial"
DARK_BLUE = "1F4E79"
MED_BLUE  = "2E75B6"
WHITE     = "FFFFFF"
ORANGE    = "FF8C00"
GREEN     = "70AD47"
PURPLE    = "7030A0"
TEAL      = "00B0F0"
GREY      = "595959"

def _side(c="CCCCCC"): return Side(style="thin", color=c)
def tbdr(c="CCCCCC"):  s=_side(c); return Border(left=s,right=s,top=s,bottom=s)

def hdr(cell, bg=DARK_BLUE, fg=WHITE, sz=10):
    cell.font      = Font(name=F, bold=True, color=fg, size=sz)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = tbdr("AAAAAA")

def sc_style(cell, sc, sfg):
    bgs = {5:"FF0000", 4:"FF8C00", 3:"FFD700", 2:"92D050", 1:"00B050"}
    cell.fill      = PatternFill("solid", fgColor=bgs[min(sc,5)])
    cell.font      = Font(name=F, bold=(sc>=4), color=sfg, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def note_row(ws, row_num, col_span, text, fg, bg, height=36):
    ws.merge_cells(f"A{row_num}:{get_column_letter(col_span)}{row_num}")
    c = ws.cell(row_num, 1, text)
    c.font      = Font(name=F, italic=True, size=8, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row_num].height = height

# ---------------------------------------------------------------------------
# Name / issuer helpers
# ---------------------------------------------------------------------------
def flip(name):
    if not name: return ""
    name = str(name).strip()
    if "," in name:
        p = name.split(",", 1)
        return f"{p[1].strip()} {p[0].strip()}"
    return name

EP_FIXES = {
    "Huang James":      "James Huang",
    "Benjmain Chung":   "Benjamin Chung",
    "Benjamin J Chung": "Benjamin Chung",
    "Chung Benjamin":   "Benjamin Chung",
}
def fx(n): return EP_FIXES.get(n.strip(), n.strip())

ISSUER_NORM = {
    "Muscle Maker, Inc.":                         "Sadot Group Inc.",
    "Sadot Group, Inc.":                          "Sadot Group Inc.",
    "Phaos Technology (Cayman) Holdings Limited": "Phaos Technology Holdings (Cayman) Ltd",
    "Maison Solutions Inc.":                      "Maison Solutions, Inc.",
    "Vyome Therapeutics Inc.":                    "Vyome Holdings, Inc.",
    "Vyome Holdings, Inc.":                       "Vyome Holdings, Inc.",
    "IMMRSIV Inc.":                               "Immrsiv Inc.",
    "Datasea, Inc.":                              "Datasea Inc.",
    "Sigyn Therapeutics, Inc.":                   "Sigyn Therapeutics, Inc",
    "VPR Brands, LP":                             "VPR Brands LP",
    "VPR Brands, LP.":                            "VPR Brands LP",
    "VPR Brands, L.P.":                           "VPR Brands LP",
    "Franklin Wireless Corp.":                    "Franklin Wireless Corp",
    "I-ON Digital Corp.":                         "I-ON Digital Corp",
    "Boomer Holdings, Inc":                       "Boomer Holdings Inc",
    "Thunder Energies Corporation":               "Thunder Energies Corp",
}
def fxi(n): n=n.replace("\xa0"," ").strip(); return ISSUER_NORM.get(n, n)

# Bizinta display name -> canonical PCAOB issuer name (None = no PCAOB filing expected)
BIZINTA_TO_PCAOB = {
    "AlphaTime Acquisition Corp.":                       None,
    "Alternus Clean Energy":                             "Alternus Clean Energy, Inc.",
    "American Senior Association Holding Group, Inc.":   None,
    "Baird Medical":                                     "Baird Medical Investment Holdings Ltd",
    "Baiya":                                             "Baiya International Group Inc.",
    "Brava Acquisition Corp":                            "Brava Acquisition Corp",
    "Btab Ecommerce Group, Inc.":                        "Btab Ecommerce Group, Inc.",
    "Concorde International Group Ltd":                  "Concorde International Group Ltd.",
    "CRK":                                               None,
    "Cuprina":                                           "Cuprina Holdings (Cayman) LTD",
    "Curanex Pharmaceuticals":                           "Curanex Pharmaceuticals Inc",
    "Datasea":                                           "Datasea Inc.",
    "Durango Gold Corp":                                 None,
    "EDETEK Inc.":                                       None,
    "Followone Inc.":                                    None,
    "Fonon Corporation":                                 None,
    "FOXO Technologies Inc.":                            "FOXO Technologies Inc.",
    "FutureCrest Acquisition Corp. II":                  None,
    "FutureCrest Acquisition Corp III":                  None,
    "Galle Technology Limited":                          None,
    "GATC Health Corp.":                                 None,
    "Geoswift Digital Group Limited":                    None,
    "G-mango Inc.":                                      None,
    "Goodvision AI Inc.":                                None,
    "Graphjet Technology":                               "Graphjet Technology",
    "HDEDUCATION":                                       None,
    "Inspire Veterinary Partners":                       "Inspire Veterinary Partners, Inc.",
    "IntelliStem":                                       None,
    "I-ON Digital Corp":                                 "I-ON Digital Corp",
    "Iveda Solutions":                                   None,
    "IWAC Holding Company Inc.":                         "IWAC Holding Co Inc.",
    "iZooto":                                            None,
    "JTS - Outsourcing.":                                None,
    "Kandi Technologies Group, Inc.":                    "Kandi Technologies Group, Inc.",
    "Kat Fabricators":                                   None,
    "Knorex":                                            "Knorex Ltd.",
    "Laser Photonics Corporation":                       None,
    "LDR":                                               None,
    "Lithium & Boron Technology, Inc.":                  "Lithium & Boron Technology, Inc.",
    "Lokahi Therapeutics (fka Apimeds Pharmaceuticals)": "Apimeds Pharmaceuticals US, Inc.",
    "Maison Solutions, Inc.":                            "Maison Solutions, Inc.",
    "Medera Inc.":                                       "Medera Inc.",
    "Meihua International Medical Technologies Co., Ltd.":"Meihua International Medical Technologies Co., Ltd.",
    "Non-Client Projects":                               None,
    "NovelStem International Corp.":                     "NovelStem International Corp.",
    "Novusterra":                                        "Novusterra Inc.",
    "Olayan America":                                    None,
    "PCAOB":                                             None,
    "Phaos":                                             "Phaos Technology Holdings (Cayman) Ltd",
    "Porche Capital SEE 1 Acquisition Corp.":            None,
    "Precision Aerospace Group Inc":                     "Precision Aerospace & Defense Group, Inc.",
    "PT Transcoal":                                      None,
    "RainRock Acquisition Corp":                         None,
    "Restake Technologies Ltd.":                         None,
    "Ryde Technology":                                   "Ryde Group Ltd",
    "Sadot Group Inc.":                                  "Sadot Group Inc.",
    "SDE":                                               None,
    "Shanghai Yihang Network Technology Co., Ltd":       None,
    "Shreya Acquisition Group":                          "Shreya Acquisition Group",
    "Sigyn Therapeutic":                                 "Sigyn Therapeutics, Inc",
    "Skubbs Pte. Ltd.":                                  "Skubbs Holdings Ltd",
    "Test Company":                                      None,
    "TJGC Group Limited (fka CTRL Media)":               "Ctrl Group Limited",
    "TrueTribe":                                         None,
    "Unicoin Inc.":                                      "Unicoin Inc.",
    "VPR Brands LP":                                     "VPR Brands LP",
    "Vyome Holdings, Inc.":                              "Vyome Holdings, Inc.",
    "Zooming (Korea) Co., Ltd.":                         None,
}
PCAOB_TO_BIZINTA = {v: k for k, v in BIZINTA_TO_PCAOB.items() if v}

EQR_PROP_ID   = "45173303"
GHOST_CLIENTS = {"Epione Health", "Fuse Group Holding Inc", "2022 Audit"}

# ---------------------------------------------------------------------------
# Data fetchers
# ---------------------------------------------------------------------------
def fetch_bizinta():
    endpoint = f"https://{BIZINTA_SUBDOMAIN}.bizinta.com/graphql"
    query = """
    {
      organizations(filters: {}) {
        nodes {
          displayName
          pipelineStatus { displayName }
          clientManager { displayName }
          genericPropValues {
            genericProp { id }
            genericPropValues { displayName }
          }
        }
      }
    }
    """
    payload = json.dumps({"query": query}).encode("utf-8")
    req = urllib.request.Request(endpoint, data=payload, headers={
        "Authorization":   f"Bearer {BIZINTA_TOKEN}",
        "Content-Type":    "application/json",
        "Accept":          "application/json",
        "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Origin":          f"https://{BIZINTA_SUBDOMAIN}.bizinta.com",
        "Referer":         f"https://{BIZINTA_SUBDOMAIN}.bizinta.com/graphql/ide",
    }, method="POST")
    with urllib.request.urlopen(req, timeout=20) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    out = {}
    for c in data["data"]["organizations"]["nodes"]:
        nm = c["displayName"]
        if nm in GHOST_CLIENTS: continue
        status = (c.get("pipelineStatus") or {}).get("displayName", "")
        ep     = flip((c.get("clientManager") or {}).get("displayName", ""))
        eqr    = ""
        for info in c.get("genericPropValues") or []:
            if (info.get("genericProp") or {}).get("id") == EQR_PROP_ID:
                vals = info.get("genericPropValues") or []
                if vals: eqr = vals[0].get("displayName", "") or ""
                break
        out[nm] = {"status": status, "ep": ep, "eqr": eqr}
    return out

def fetch_pcaob():
    with urllib.request.urlopen(FILTER_SERVICE_URL, timeout=60) as resp:
        raw = resp.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(raw))
    all_bulk, raw_rows = [], []
    for row in reader:
        raw_rows.append(dict(row))
        if row.get("Latest Form AP Filing", "").strip() != "1": continue
        fn = row.get("Engagement Partner First Name", "").strip()
        mn = row.get("Engagement Partner Middle Name", "").strip()
        ln = row.get("Engagement Partner Last Name", "").strip()
        ep_full   = fx((f"{fn} {mn} {ln}" if mn else f"{fn} {ln}").strip())
        fpe_raw   = row.get("Fiscal Period End Date", "").strip().split()[0]
        try:
            fpe_dt  = _dt.strptime(fpe_raw, "%m/%d/%Y")
            fye_str = fpe_dt.strftime("%Y-%m-%d")
            yr      = fpe_dt.year
        except:
            fye_str = fpe_raw; yr = 0
        filed_raw = row.get("Filing Date", "").strip().split()[0]
        try:    filed = _dt.strptime(filed_raw, "%m/%d/%Y").strftime("%Y-%m-%d")
        except: filed = filed_raw
        signer     = fx(f"{row.get('Signed First Name','').strip()} {row.get('Signed Last Name','').strip()}".strip())
        issuer_raw = row.get("Issuer Name", "")
        issuer_id  = row.get("Issuer CIK", "").strip()
        firm_name  = row.get("Firm Name", "").strip()
        all_bulk.append({
            "year": yr, "ep": ep_full,
            "issuer":     fxi(issuer_raw),
            "issuer_raw": issuer_raw.strip(),
            "issuer_id":  issuer_id,
            "firm_name":  firm_name,
            "signer": signer, "filed": filed, "fye": fye_str,
        })
    groups = defaultdict(list)
    for r in all_bulk:
        groups[(r["ep"], r["issuer"], r["fye"])].append(r)
    records = sorted(
        [max(g, key=lambda x: x["filed"]) for g in groups.values()],
        key=lambda r: (r["ep"], r["issuer"], r["year"])
    )
    return records, raw_rows

# ---------------------------------------------------------------------------
# Gap detection
# ---------------------------------------------------------------------------
def detect_gap_years(enriched):
    group_years = defaultdict(set)
    for r in enriched: group_years[(r["ep"], r["issuer"])].add(r["year"])
    gap_flags = set()
    for (ep, issuer), years in group_years.items():
        sy = sorted(years)
        for i, yr in enumerate(sy):
            if i == 0: continue
            if yr - sy[i-1] > 1:
                for fy in sy[i:]: gap_flags.add((ep, issuer, fy))
    return gap_flags

# ---------------------------------------------------------------------------
# Rotation logic
# ---------------------------------------------------------------------------
def build_rotation(records, bizinta_data):
    """
    Builds enriched (all filings) and dashboard (latest per EP-issuer).
    dashboard is filtered to active Bizinta clients only.
    """
    group_years = defaultdict(set)
    for r in records: group_years[(r["ep"], r["issuer"])].add(r["year"])

    def consec(ep, issuer, yr):
        ys = group_years[(ep, issuer)]; c = 1; y = yr - 1
        while y in ys: c += 1; y -= 1
        return c

    def rot_status(c):
        if c >= 5: return (5, "CRITICAL - Year 5+ (Rotate Now)", "FF0000", "FFFFFF")
        if c == 4: return (4, "WARNING - Year 4 (Plan Rotation)", "FF8C00", "FFFFFF")
        if c == 3: return (3, "MONITOR - Year 3",                 "FFD700", "000000")
        if c == 2: return (2, "OK - Year 2",                      "92D050", "000000")
        return          (1, "OK - Year 1",                        "00B050", "FFFFFF")

    # Build set of PCAOB issuer names that are currently Active in Bizinta
    active_pcaob_issuers = set()
    for biz_name, biz_info in bizinta_data.items():
        if biz_info.get("status") == "Active":
            mapped = BIZINTA_TO_PCAOB.get(biz_name)
            if mapped: active_pcaob_issuers.add(mapped)

    enriched = []
    for r in records:
        c = consec(r["ep"], r["issuer"], r["year"])
        sc, sl, sbg, sfg = rot_status(c)
        all_yrs  = sorted(group_years[(r["ep"], r["issuer"])])
        start_yr = r["year"] - c + 1
        consec_chain = []
        y = r["year"]
        while y in group_years[(r["ep"], r["issuer"])]:
            consec_chain.append(y); y -= 1
        consec_chain = sorted(consec_chain)
        calc_note = (
            f"Years on file: {', '.join(str(y) for y in all_yrs)}. "
            f"Consecutive chain ending {r['year']}: {' > '.join(str(y) for y in consec_chain)} = {c} year(s). "
            f"Rotation limit: 5 years. Years remaining: {max(0, 5-c)}."
        )
        if len(all_yrs) > len(consec_chain):
            gaps = [f"{all_yrs[i-1]}-{all_yrs[i]}" for i in range(1,len(all_yrs)) if all_yrs[i]-all_yrs[i-1]>1]
            calc_note += f" NOTE: Gap(s) detected ({', '.join(gaps)}) - count restarted after gap."
        enriched.append({**r,
            "consec": c, "sc": sc, "sl": sl, "sbg": sbg, "sfg": sfg,
            "yrs_left": max(0, 5-c), "start_yr": start_yr,
            "all_yrs": all_yrs, "calc_note": calc_note,
        })

    # Dashboard: latest filing per EP-issuer, restricted to active Bizinta clients
    latest_map = {}
    for r in enriched:
        k = (r["ep"], r["issuer"])
        if k not in latest_map or r["year"] > latest_map[k]["year"]:
            latest_map[k] = r
    dashboard = sorted(
        [r for r in latest_map.values() if r["issuer"] in active_pcaob_issuers],
        key=lambda r: (-r["sc"], r["ep"], r["issuer"])
    )
    return enriched, dashboard

# ---------------------------------------------------------------------------
# Bizinta status lookup helpers
# ---------------------------------------------------------------------------
def biz_status_for_pcaob(issuer_norm, bizinta_data):
    """Returns (bizinta_name, status_label) for a PCAOB issuer name."""
    biz_name = PCAOB_TO_BIZINTA.get(issuer_norm, "")
    if biz_name and biz_name in bizinta_data:
        raw = bizinta_data[biz_name].get("status", "")
        if raw == "Active":   return biz_name, "Active in Bizinta"
        if raw == "Inactive": return biz_name, "Inactive in Bizinta"
        return biz_name, f"In Bizinta ({raw})"
    if biz_name: return biz_name, "In Bizinta (status unknown)"
    return "", "Not found in Bizinta"

def biz_ep_for_pcaob(issuer_norm, bizinta_data):
    biz_name = PCAOB_TO_BIZINTA.get(issuer_norm, "")
    if biz_name and biz_name in bizinta_data:
        return bizinta_data[biz_name].get("ep", "")
    return ""

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------
def build_excel(enriched, dashboard, pcaob_to_eqr, bizinta_data, raw_pcaob_rows, run_date_str):
    wb    = Workbook()
    gf    = detect_gap_years(enriched)
    NCOLS = 14  # max columns across sheets

    # Set of PCAOB issuers that are active in Bizinta (for reconciliation)
    active_pcaob = set()
    inactive_pcaob = set()
    for biz_name, info in bizinta_data.items():
        mapped = BIZINTA_TO_PCAOB.get(biz_name)
        if not mapped: continue
        if info.get("status") == "Active":   active_pcaob.add(mapped)
        if info.get("status") == "Inactive": inactive_pcaob.add(mapped)
    all_pcaob_issuers = set(r["issuer"] for r in enriched)

    # -----------------------------------------------------------------------
    # Sheet 1: Rotation Dashboard (active Bizinta clients only)
    # -----------------------------------------------------------------------
    ws1 = wb.active; ws1.title = "Rotation Dashboard"
    ws1.sheet_properties.tabColor = DARK_BLUE

    ws1.merge_cells("A1:N1")
    ws1["A1"] = "KREIT & CHIU CPA LLP - PCAOB Partner Rotation Tracker (AS 1201)"
    ws1["A1"].font      = Font(name=F, bold=True, size=14, color=DARK_BLUE)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:N2")
    ws1["A2"] = (
        f"Source: PCAOB Form AP Filings (Firm ID 6651 - all name variants) | "
        f"EQR: Bizinta API (live) | Generated: {run_date_str} | "
        f"{len(enriched)} total filings | {len(dashboard)} active engagements "
        f"(active Bizinta clients with PCAOB filings only)"
    )
    ws1["A2"].font      = Font(name=F, italic=True, size=9, color=GREY)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 16

    # Scope assumption note (per review + confirmed interpretation)
    note_row(ws1, 3, 14,
        "SCOPE: This dashboard includes ONLY clients that are (1) currently Active in Bizinta "
        "AND (2) have at least one Form AP filing on record with PCAOB under Firm ID 6651. "
        "Clients active in Bizinta with no PCAOB filing are listed in the 'Active Bizinta - No Filing' tab. "
        "Clients with PCAOB filings but no longer active in Bizinta are excluded - the partner rotation "
        "chain is considered broken once the firm is no longer engaged.",
        "1F4E79", "DEEAF1", height=40)

    note_row(ws1, 4, 14,
        "EQR ASSUMPTION (B - CRITICAL): EQR tenure assumed to start same year as EP. "
        "PCAOB Form AP does not record EQR appointment dates. "
        "Human review required where the firm holds internal records showing a different EQR start date. "
        "EQR is sourced from Bizinta live API (Client EQR custom field, Property ID 45173303).",
        "7F3F00", "FFF2CC")

    note_row(ws1, 5, 14,
        "GAP-YEAR ASSUMPTION (A - CRITICAL): A year with no Form AP filing breaks the consecutive chain "
        "and resets the count to Year 1 on return (literal interpretation of SEC Rule 2-01(c)(6)). "
        "Rows marked [GAP] indicate a gap-then-return pattern - human review required. "
        "An engineered gap to reset the clock constitutes circumvention under PCAOB rules.",
        "7F0000", "FFE4E1")

    DASH_HDRS = [
        "EP as per\nForm AP Filing",
        "EP as per\nBizinta",
        "Issuer / Client",
        "EP Start\nYear",
        "Latest\nAudit Yr",
        "Consec.\nYears",
        "Yrs\nLeft",
        "Rotation Status",
        "Signer\n(Form AP)",
        "EQR\n(Bizinta)",
        "Fiscal\nYear End",
        "Last Filed",
        "Rotation\nDeadline",
        "Gap\nFlag",
    ]
    ws1.row_dimensions[6].height = 44
    for c, h in enumerate(DASH_HDRS, 1): hdr(ws1.cell(6, c, h))
    CW1 = [22, 22, 34, 11, 12, 11, 10, 30, 20, 24, 13, 13, 14, 9]
    for i, w in enumerate(CW1, 1): ws1.column_dimensions[get_column_letter(i)].width = w

    for ri, r in enumerate(dashboard, 7):
        eqr_val   = pcaob_to_eqr.get(r["issuer"], "")
        biz_ep    = biz_ep_for_pcaob(r["issuer"], bizinta_data)
        left_str  = "ROTATE NOW" if r["yrs_left"] == 0 else str(r["yrs_left"])
        fye       = r["fye"][:10] if r["fye"] else ""
        deadline  = str(r["start_yr"] + 4) if r["start_yr"] else ""
        has_gap   = any((r["ep"], r["issuer"], yr) in gf for yr in r["all_yrs"])
        gap_flag  = "[GAP]" if has_gap else ""

        row_vals = [r["ep"], biz_ep, r["issuer"], r["start_yr"], r["year"],
                    r["consec"], left_str, r["sl"], r["signer"], eqr_val,
                    fye, r["filed"][:10], deadline, gap_flag]
        for ci, v in enumerate(row_vals, 1):
            cell = ws1.cell(ri, ci, v)
            cell.font      = Font(name=F, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in [3, 8]))
            cell.border    = tbdr("DDDDDD")
            if ci in (6, 8):
                sc_style(cell, r["sc"], r["sfg"])
            elif ci == 7:
                if r["yrs_left"] == 0:
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                    cell.font = Font(name=F, bold=True, color=WHITE, size=10)
                elif r["yrs_left"] == 1:
                    cell.fill = PatternFill("solid", fgColor=ORANGE)
                    cell.font = Font(name=F, bold=True, color=WHITE, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 4:
                cell.font  = Font(name=F, bold=True, size=11, color=DARK_BLUE)
                cell.fill  = PatternFill("solid", fgColor="DEEAF1")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 2 and biz_ep:
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
                cell.font = Font(name=F, size=10, color="375623")
            elif ci == 10:
                bg_eqr = "E2EFDA" if eqr_val else "FFF2CC"
                cell.fill = PatternFill("solid", fgColor=bg_eqr)
                if not eqr_val:
                    cell.font  = Font(name=F, size=9, italic=True, color="AA6600")
                    cell.value = "(enter EQR)"
            elif ci == 14 and has_gap:
                cell.fill  = PatternFill("solid", fgColor=ORANGE)
                cell.font  = Font(name=F, bold=True, color=WHITE, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[ri].height = 18
    ws1.freeze_panes = "A7"

    # -----------------------------------------------------------------------
    # Sheet 2: PCAOB Filings with Bizinta Status (Section 2.2)
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("PCAOB Filings - Biz Status")
    ws2.sheet_properties.tabColor = MED_BLUE

    ws2.merge_cells("A1:O1")
    ws2["A1"] = (f"PCAOB FORM AP FILINGS - Firm ID 6651 (all name variants) | "
                 f"Bizinta Status matched | {len(enriched)} records | {run_date_str}")
    ws2["A1"].font      = Font(name=F, bold=True, size=12, color=DARK_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    note_row(ws2, 2, 15,
        "Bizinta status column shows whether the PCAOB-filed client is currently Active, Inactive, or Not Found in Bizinta. "
        "Filtering by Firm ID 6651 captures all filings regardless of firm name variant "
        "(Benjamin & Co, Kreit & Chiu CPA LLP, Paris Kreit & Chiu CPA LLP).",
        DARK_BLUE, "DEEAF1", height=28)

    F2_HDRS = [
        "Audit\nYear", "Filing Date", "Firm Name\n(as filed)",
        "EP as per\nForm AP", "EP as per\nBizinta",
        "Issuer / Client (Normalised)", "Issuer CIK",
        "PCAOB Raw Name", "Consec.\nYear #",
        "Rotation Status", "Signer (Form AP)",
        "EQR (Bizinta)", "Fiscal Year End",
        "Yrs Left", "Bizinta Status",
    ]
    ws2.row_dimensions[3].height = 40
    for c, h in enumerate(F2_HDRS, 1): hdr(ws2.cell(3, c, h), bg=MED_BLUE)
    CW2 = [9, 13, 22, 22, 22, 36, 12, 36, 11, 28, 20, 22, 14, 9, 20]
    for i, w in enumerate(CW2, 1): ws2.column_dimensions[get_column_letter(i)].width = w

    for ri, r in enumerate(enriched, 4):
        eqr_v     = pcaob_to_eqr.get(r["issuer"], "")
        biz_ep    = biz_ep_for_pcaob(r["issuer"], bizinta_data)
        _, biz_st = biz_status_for_pcaob(r["issuer"], bizinta_data)
        has_gap   = (r["ep"], r["issuer"], r["year"]) in gf
        rv = [r["year"], r["filed"][:10], r.get("firm_name",""),
              r["ep"], biz_ep,
              r["issuer"], r.get("issuer_id",""), r.get("issuer_raw", r["issuer"]),
              r["consec"], r["sl"], r["signer"],
              eqr_v, r["fye"][:10] if r["fye"] else "",
              r["yrs_left"], biz_st]
        for ci, v in enumerate(rv, 1):
            cell = ws2.cell(ri, ci, v)
            cell.font      = Font(name=F, size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in [6, 8, 10]))
            cell.border    = tbdr("EEEEEE")
            if ci in (9, 10):
                sc_style(cell, r["sc"], r["sfg"])
            elif ci == 15:
                if "Active"   in biz_st: cell.fill = PatternFill("solid", fgColor="E2EFDA"); cell.font = Font(name=F, size=9, color="375623")
                elif "Inactive" in biz_st: cell.fill = PatternFill("solid", fgColor="FFF2CC"); cell.font = Font(name=F, size=9, color="7F3F00")
                else: cell.fill = PatternFill("solid", fgColor="FCE4D6"); cell.font = Font(name=F, size=9, color="7F0000")
            if has_gap and ci not in (9, 10, 15):
                if not (cell.fill.patternType == "solid"):
                    cell.fill = PatternFill("solid", fgColor="FFF0E0")
        ws2.row_dimensions[ri].height = 15
    ws2.freeze_panes = "A4"

    # -----------------------------------------------------------------------
    # Sheet 3: Active Bizinta Clients with No PCAOB Filing (Section 2.3)
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Active Bizinta - No Filing")
    ws3.sheet_properties.tabColor = ORANGE

    # Build list: all active Bizinta clients with no matching PCAOB issuer
    no_filing = []
    for biz_name, info in sorted(bizinta_data.items()):
        if info.get("status") != "Active": continue
        mapped = BIZINTA_TO_PCAOB.get(biz_name)
        # Covered = mapped PCAOB name exists AND has at least one filing
        if mapped and mapped in all_pcaob_issuers:
            continue
        reason = "No PCAOB filing found" if mapped else "No PCAOB mapping defined - verify if PCAOB client"
        no_filing.append({
            "biz_name":   biz_name,
            "pcaob_name": mapped or "",
            "ep":         info.get("ep", ""),
            "eqr":        info.get("eqr", ""),
            "reason":     reason,
        })

    ws3.merge_cells("A1:E1")
    ws3["A1"] = f"ACTIVE BIZINTA CLIENTS WITH NO PCAOB FILING FOUND | {len(no_filing)} clients | {run_date_str}"
    ws3["A1"].font      = Font(name=F, bold=True, size=12, color=DARK_BLUE)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    note_row(ws3, 2, 5,
        "These clients are currently Active in Bizinta but have no corresponding Form AP filing "
        "found in the PCAOB database under Firm ID 6651. They may be non-PCAOB clients, newly onboarded clients, "
        "or clients where the PCAOB name mapping has not yet been defined. "
        "Manual review is required to determine whether a PCAOB filing is expected.",
        "7F3F00", "FFF2CC", height=40)

    NF_HDRS = ["Bizinta Client Name", "PCAOB Mapped Name", "EP (Bizinta)", "EQR (Bizinta)", "Status / Notes"]
    ws3.row_dimensions[3].height = 32
    for c, h in enumerate(NF_HDRS, 1): hdr(ws3.cell(3, c, h), bg=ORANGE)
    CW3 = [38, 38, 24, 24, 42]
    for i, w in enumerate(CW3, 1): ws3.column_dimensions[get_column_letter(i)].width = w

    for ri, item in enumerate(no_filing, 4):
        for ci, v in enumerate([item["biz_name"], item["pcaob_name"],
                                 item["ep"], item["eqr"], item["reason"]], 1):
            cell = ws3.cell(ri, ci, v)
            cell.font = Font(name=F, size=10)
            cell.border = tbdr("DDDDDD")
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in [1, 2, 5]))
            if not item["pcaob_name"]:
                cell.fill = PatternFill("solid", fgColor="FFF0E0")
            else:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
        ws3.row_dimensions[ri].height = 17

    # -----------------------------------------------------------------------
    # Sheet 4: Partner Summary (rebuilt from corrected scope)
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Partner Summary")
    ws4.sheet_properties.tabColor = GREEN

    ws4.merge_cells("A1:J1")
    ws4["A1"] = "ENGAGEMENT PARTNER ACTIVITY SUMMARY (Active Bizinta clients scope)"
    ws4["A1"].font = Font(name=F, bold=True, size=13, color=DARK_BLUE)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    active_eps  = set()
    for biz_name, info in bizinta_data.items():
        if info.get("status") == "Active" and info.get("ep"):
            active_eps.add(info["ep"])

    pdata = defaultdict(lambda: {"engagements":set(), "years":set(), "max_c":0,
                                  "crit":0, "warn":0, "mon":0, "is_active":False})
    for r in dashboard:
        p = pdata[r["ep"]]
        p["engagements"].add(r["issuer"])
        p["years"].add(r["year"])
        p["max_c"] = max(p["max_c"], r["consec"])
        if r["sc"] >= 5: p["crit"] += 1
        elif r["sc"] == 4: p["warn"] += 1
        elif r["sc"] == 3: p["mon"] += 1
        if r["ep"] in active_eps: p["is_active"] = True

    SUM_HDRS = ["Engagement Partner", "Form AP EP", "Status",
                "Active Clients\n(Dashboard)", "Years Active\n(Range)",
                "Max Consec.\nYears", "Critical\n(Yr 5+)",
                "Warning\n(Yr 4)", "Monitor\n(Yr 3)", "Overall Risk"]
    CW4 = [24, 24, 12, 16, 16, 16, 14, 14, 14, 16]
    for i, w in enumerate(CW4, 1): ws4.column_dimensions[get_column_letter(i)].width = w

    def write_partner_section(ws, start_row, title, bg, partners):
        ws.merge_cells(f"A{start_row}:J{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(name=F, bold=True, size=11, color=WHITE)
        ws[f"A{start_row}"].fill = PatternFill("solid", fgColor=bg)
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 22
        hr = start_row + 1
        for c, h in enumerate(SUM_HDRS, 1): hdr(ws.cell(hr, c, h), bg=bg)
        ws.row_dimensions[hr].height = 40
        cr = hr + 1
        for ep, pd in sorted(partners, key=lambda x: (-x[1]["crit"], -x[1]["warn"], x[0])):
            yrs  = sorted(pd["years"])
            risk = ("HIGH RISK","FF0000","FFFFFF") if pd["crit"] else \
                   ("MEDIUM RISK","FF8C00","FFFFFF") if pd["warn"] else \
                   ("MONITOR","FFD700","000000") if pd["mon"] else ("LOW RISK","00B050","FFFFFF")
            biz_ep_name = ep  # EP here is from dashboard (Form AP EP)
            rv = [biz_ep_name, ep, "Active" if pd["is_active"] else "Departed",
                  len(pd["engagements"]),
                  f"{yrs[0]}-{yrs[-1]}" if yrs else "",
                  pd["max_c"], pd["crit"], pd["warn"], pd["mon"], risk[0]]
            for ci, v in enumerate(rv, 1):
                cell = ws.cell(cr, ci, v)
                cell.font = Font(name=F, size=10)
                cell.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center")
                cell.border = tbdr()
                if ci == 3:
                    cell.fill = PatternFill("solid", fgColor="375623" if v=="Active" else GREY)
                    cell.font = Font(name=F, bold=True, color=WHITE, size=9)
                elif ci == 10:
                    cell.fill = PatternFill("solid", fgColor=risk[1])
                    cell.font = Font(name=F, bold=True, color=risk[2], size=10)
                elif ci == 6:
                    mc = pd["max_c"]
                    sc_style(cell, min(mc, 5), WHITE if mc>=3 else "000000")
            ws.row_dimensions[cr].height = 20
            cr += 1
        return cr

    active_list   = [(ep, pd) for ep, pd in pdata.items() if pd["is_active"]]
    inactive_list = [(ep, pd) for ep, pd in pdata.items() if not pd["is_active"]]
    cr = write_partner_section(ws4, 2, "SECTION A: ACTIVE PARTNERS", "375623", active_list)
    cr += 1
    write_partner_section(ws4, cr, "SECTION B: DEPARTED / INACTIVE PARTNERS", GREY, inactive_list)

    # -----------------------------------------------------------------------
    # Sheet 5: Client Reconciliation (corrected logic)
    # -----------------------------------------------------------------------
    ws5 = wb.create_sheet("Client Reconciliation")
    ws5.sheet_properties.tabColor = PURPLE

    ws5.merge_cells("A1:F1")
    ws5["A1"] = "CLIENT RECONCILIATION - Three-way match: PCAOB Filings vs Bizinta Active Clients"
    ws5["A1"].font = Font(name=F, bold=True, size=13, color=DARK_BLUE)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28
    CW5 = [40, 40, 18, 24, 24, 36]
    for i, w in enumerate(CW5, 1): ws5.column_dimensions[get_column_letter(i)].width = w

    def recon_section(ws, ri, title, bg, hdrs, rows_fn):
        ws.merge_cells(f"A{ri}:F{ri}")
        ws[f"A{ri}"] = title
        ws[f"A{ri}"].fill = PatternFill("solid", fgColor=bg)
        ws[f"A{ri}"].font = Font(name=F, bold=True, size=11, color=WHITE)
        ws[f"A{ri}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[ri].height = 22
        ri += 1
        for c, h in enumerate(hdrs, 1): hdr(ws.cell(ri, c, h), bg=bg)
        ws.row_dimensions[ri].height = 32
        ri += 1
        ri = rows_fn(ws, ri)
        return ri + 1

    # Category 1: PCAOB filing AND active in Bizinta
    cat1 = sorted([iss for iss in all_pcaob_issuers if iss in active_pcaob])

    def cat1_rows(ws, ri):
        for issuer in cat1:
            biz_n = PCAOB_TO_BIZINTA.get(issuer, "")
            last_ep, last_yr, last_biz_ep = "", 0, ""
            for r in enriched:
                if r["issuer"] == issuer and r["year"] > last_yr:
                    last_yr = r["year"]; last_ep = r["ep"]
            if biz_n and biz_n in bizinta_data:
                last_biz_ep = bizinta_data[biz_n].get("ep", "")
            for ci, v in enumerate([issuer, biz_n, "Active in Bizinta", last_ep, last_biz_ep, ""], 1):
                cell = ws.cell(ri, ci, v)
                cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
                cell.alignment = Alignment(vertical="center", wrap_text=(ci in [1,2,6]))
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
            ws.row_dimensions[ri].height = 16; ri += 1
        return ri

    # Category 2: PCAOB filing but NOT found/active in Bizinta
    cat2 = sorted([iss for iss in all_pcaob_issuers if iss not in active_pcaob])

    def cat2_rows(ws, ri):
        for issuer in cat2:
            biz_n, biz_st = biz_status_for_pcaob(issuer, bizinta_data)
            last_ep, last_yr = "", 0
            for r in enriched:
                if r["issuer"] == issuer and r["year"] > last_yr:
                    last_yr = r["year"]; last_ep = r["ep"]
            note = biz_st
            for ci, v in enumerate([issuer, biz_n, biz_st, last_ep, "", note], 1):
                cell = ws.cell(ri, ci, v)
                cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
                cell.alignment = Alignment(vertical="center", wrap_text=(ci in [1,2,6]))
                if "Inactive" in biz_st: cell.fill = PatternFill("solid", fgColor="FFF2CC")
                else: cell.fill = PatternFill("solid", fgColor="FCE4D6")
            ws.row_dimensions[ri].height = 16; ri += 1
        return ri

    # Category 3: Active in Bizinta but no PCAOB filing
    def cat3_rows(ws, ri):
        for item in no_filing:
            for ci, v in enumerate([item["pcaob_name"] or "(no mapping)", item["biz_name"],
                                     "No PCAOB filing found", item["ep"], item["eqr"], item["reason"]], 1):
                cell = ws.cell(ri, ci, v)
                cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
                cell.alignment = Alignment(vertical="center", wrap_text=(ci in [1,2,6]))
                cell.fill = PatternFill("solid", fgColor="FFF0E0")
            ws.row_dimensions[ri].height = 16; ri += 1
        return ri

    HDR_CAT = ["PCAOB Issuer Name", "Bizinta Client Name", "Bizinta Status",
               "Last EP (Form AP)", "EP (Bizinta)", "Notes"]
    ri = 2
    ri = recon_section(ws5, ri,
        f"CATEGORY 1: PCAOB filing client ALSO active in Bizinta ({len(cat1)})",
        "375623", HDR_CAT, cat1_rows)
    ri = recon_section(ws5, ri,
        f"CATEGORY 2: PCAOB filing client NOT active/found in Bizinta ({len(cat2)})",
        DARK_BLUE, HDR_CAT, cat2_rows)
    ri = recon_section(ws5, ri,
        f"CATEGORY 3: Active Bizinta client with NO PCAOB filing ({len(no_filing)})",
        ORANGE, HDR_CAT, cat3_rows)

    # -----------------------------------------------------------------------
    # Sheet 6: Name Changes (3-category structure per Section 7)
    # -----------------------------------------------------------------------
    ws6 = wb.create_sheet("Name Changes")
    ws6.sheet_properties.tabColor = TEAL
    ws6.merge_cells("A1:D1")
    ws6["A1"] = "NAME CHANGES & NORMALISATION - PCAOB Issuer ID as primary key"
    ws6["A1"].font = Font(name=F, bold=True, size=13, color=DARK_BLUE)
    ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 28
    note_row(ws6, 2, 4,
        "Three categories: (1) Appears in PCAOB AND active in Bizinta. "
        "(2) Appears in PCAOB but NOT found in Bizinta - check for name mismatch. "
        "(3) Active in Bizinta but NOT in PCAOB filings - may be non-issuer or new client. "
        "Issuer CIK is the authoritative key per PCAOB records.",
        DARK_BLUE, "DEEAF1", height=36)
    NC_HDRS = ["Issuer CIK", "PCAOB Raw Name (as filed)", "Normalised Name (tracker)", "Bizinta Name / Status"]
    CW6 = [16, 46, 46, 40]
    for i, w in enumerate(CW6, 1): ws6.column_dimensions[get_column_letter(i)].width = w

    def nc_section(ws, ri, title, bg):
        ws.merge_cells(f"A{ri}:D{ri}")
        ws[f"A{ri}"] = title
        ws[f"A{ri}"].fill = PatternFill("solid", fgColor=bg)
        ws[f"A{ri}"].font = Font(name=F, bold=True, size=10, color=WHITE)
        ws[f"A{ri}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[ri].height = 20; ri += 1
        for c, h in enumerate(NC_HDRS, 1): hdr(ws.cell(ri, c, h), bg=bg)
        ws.row_dimensions[ri].height = 32; ri += 1
        return ri

    nc_ri = nc_section(ws6, 3, "CATEGORY 1: PCAOB filing client active in Bizinta", "375623")
    shown = set()
    for r in enriched:
        if r["issuer"] not in active_pcaob: continue
        key = (r.get("issuer_id",""), r.get("issuer_raw", r["issuer"]))
        if key in shown: continue
        shown.add(key)
        biz_n = PCAOB_TO_BIZINTA.get(r["issuer"], "")
        biz_st = bizinta_data.get(biz_n, {}).get("status", "") if biz_n else ""
        label  = f"{biz_n} ({biz_st})" if biz_n else "(no Bizinta mapping)"
        bg = "E2EFDA" if biz_st == "Active" else "FFFFFF"
        for ci, v in enumerate([r.get("issuer_id",""), r.get("issuer_raw",r["issuer"]), r["issuer"], label], 1):
            cell = ws6.cell(nc_ri, ci, v)
            cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
            cell.alignment = Alignment(vertical="center", wrap_text=(ci>1))
            cell.fill = PatternFill("solid", fgColor=bg)
        ws6.row_dimensions[nc_ri].height = 15; nc_ri += 1

    nc_ri += 1
    nc_ri = nc_section(ws6, nc_ri, "CATEGORY 2: PCAOB filing client NOT found in Bizinta", DARK_BLUE)
    shown2 = set()
    for r in enriched:
        if r["issuer"] in active_pcaob: continue
        key = (r.get("issuer_id",""), r.get("issuer_raw", r["issuer"]))
        if key in shown2: continue
        shown2.add(key)
        biz_n, biz_st_label = biz_status_for_pcaob(r["issuer"], bizinta_data)
        for ci, v in enumerate([r.get("issuer_id",""), r.get("issuer_raw",r["issuer"]), r["issuer"], biz_st_label], 1):
            cell = ws6.cell(nc_ri, ci, v)
            cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
            cell.alignment = Alignment(vertical="center", wrap_text=(ci>1))
            cell.fill = PatternFill("solid", fgColor="FCE4D6")
        ws6.row_dimensions[nc_ri].height = 15; nc_ri += 1

    nc_ri += 1
    nc_ri = nc_section(ws6, nc_ri, "CATEGORY 3: Active Bizinta client NOT in PCAOB filings", ORANGE)
    for item in no_filing:
        for ci, v in enumerate(["", item["biz_name"], item["pcaob_name"] or "(no mapping)", "Active in Bizinta - no filing"], 1):
            cell = ws6.cell(nc_ri, ci, v)
            cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
            cell.alignment = Alignment(vertical="center", wrap_text=(ci>1))
            cell.fill = PatternFill("solid", fgColor="FFF0E0")
        ws6.row_dimensions[nc_ri].height = 15; nc_ri += 1

    # Also append static ISSUER_NORM dictionary
    nc_ri += 1
    ws6.merge_cells(f"A{nc_ri}:D{nc_ri}")
    ws6[f"A{nc_ri}"] = "STATIC NORMALISATION DICTIONARY (ISSUER_NORM) - corporate rebrands and variant spellings"
    ws6[f"A{nc_ri}"].font = Font(name=F, bold=True, size=10, color=WHITE)
    ws6[f"A{nc_ri}"].fill = PatternFill("solid", fgColor=TEAL)
    ws6[f"A{nc_ri}"].alignment = Alignment(horizontal="left", vertical="center")
    ws6.row_dimensions[nc_ri].height = 20; nc_ri += 1
    for raw_key, canonical in ISSUER_NORM.items():
        if raw_key == canonical: continue
        for ci, v in enumerate(["(varies)", raw_key, canonical, "See mapping"], 1):
            cell = ws6.cell(nc_ri, ci, v)
            cell.font = Font(name=F, size=10); cell.border = tbdr("DDDDDD")
            cell.fill = PatternFill("solid", fgColor="E8F5E9")
            cell.alignment = Alignment(vertical="center")
        ws6.row_dimensions[nc_ri].height = 15; nc_ri += 1

    # -----------------------------------------------------------------------
    # Sheet 7: Legend & Notes (updated per Section 8)
    # -----------------------------------------------------------------------
    ws7 = wb.create_sheet("Legend & Notes")
    ws7.sheet_properties.tabColor = GREY
    ws7.merge_cells("A1:C1")
    ws7["A1"] = "LEGEND, ASSUMPTIONS & METHODOLOGY NOTES"
    ws7["A1"].font = Font(name=F, bold=True, size=13, color=DARK_BLUE)
    ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[1].height = 28
    ws7.column_dimensions["A"].width = 38
    ws7.column_dimensions["B"].width = 110
    ws7.column_dimensions["C"].width = 20

    NOTES = [
        ("Dashboard Scope Rule",
         "The Rotation Dashboard includes ONLY clients that are (1) currently Active in Bizinta "
         "AND (2) have at least one Form AP filing on record with PCAOB under Firm ID 6651. "
         "Clients active in Bizinta with no PCAOB filing are listed in the 'Active Bizinta - No Filing' tab. "
         "Clients with PCAOB filings but no longer active in Bizinta are EXCLUDED from the dashboard - "
         "the partner rotation chain is considered broken once the firm is no longer engaged with that client."),
        ("PCAOB Firm ID Filter",
         "All PCAOB Form AP data is filtered by Firm ID 6651, not by firm name. "
         "This captures filings under all firm name variants: Benjamin & Co, "
         "Kreit & Chiu CPA LLP, and Paris Kreit & Chiu CPA LLP. "
         "Using firm name alone would exclude historical filings under earlier firm names."),
        ("PCAOB AS 1201 - 5 Year Rule",
         "Both the Lead Engagement Partner (EP) and the Engagement Quality Reviewer (EQR) must rotate "
         "after 5 consecutive years of service on a public company audit engagement. "
         "A 5-year cooling-off period applies after mandatory rotation before the partner may return to that engagement. "
         "Reference: SEC Rule 2-01(c)(6) and PCAOB AS 1201."),
        ("Cooling-Off Period Tracking",
         "Cooling-off period tracking requires separate logic and should be developed as a separate enhancement. "
         "The current report focuses on identifying consecutive service years based on available "
         "Form AP filing data and Bizinta information. "
         "The automation does NOT currently track whether a rotated partner has completed the "
         "5-year cooling-off period before returning to an engagement."),
        ("Consecutive Year Count",
         "One year is counted for each fiscal year end where a Form AP filing exists for the engagement. "
         "The count resets to 1 if there is a gap year with no filing (Assumption A). "
         "Example: filings in 2021, 2022, 2024 (gap in 2023) - the 2024 count is 1, not 3, "
         "because the consecutive chain was broken."),
        ("Deduplication Key - EP + Issuer Normalised + Fiscal Period End",
         "Why this key: Each Form AP filing is uniquely identified by the combination of "
         "Engagement Partner, the normalised issuer name, and the fiscal period end date. "
         "Issuer normalisation is needed because the same company can appear under different names "
         "across filings (e.g. 'Muscle Maker, Inc.' and 'Sadot Group Inc.' are the same entity after rebrand). "
         "Fiscal period end is included because some issuers have multiple fiscal year ends in the same "
         "calendar year (e.g. Brava Acquisition Corp filed for both Sep-30 and Dec-31 fiscal periods in 2025). "
         "Without fiscal period end in the key, those two filings would be collapsed into one. "
         "Duplicate Form AP records for the same key are resolved by keeping the most recently filed version. "
         "Example dedup: if two records exist for (James Huang, Datasea Inc., 2024-12-31), "
         "only the one with the later Filing Date is retained."),
        ("EQR Assumption (B - CRITICAL)",
         "PCAOB Form AP does not record EQR appointment dates. "
         "This tracker assumes EQR tenure starts at the same time as the EP (i.e. same start year). "
         "EQR data is sourced from Bizinta live API using custom field Property ID 45173303. "
         "Human review is required where the firm holds internal records showing a different EQR start date."),
        ("Gap-Year Assumption (A - CRITICAL)",
         "A year with no Form AP filing breaks the consecutive chain and resets the count to Year 1 on return. "
         "This is the literal interpretation of 'consecutive' under SEC Rule 2-01(c)(6). "
         "However, an engineered gap (deliberately rotating for one year to reset the clock) "
         "would be viewed as circumvention by PCAOB. "
         "Rows marked [GAP] indicate a gap-then-return pattern and require human review."),
        ("Ghost / Deleted Clients",
         "The following clients exist in the Bizinta API/database but do not appear in the Bizinta UI "
         "and are therefore excluded from the tracker: Epione Health, Fuse Group Holding Inc, 2022 Audit. "
         "These are believed to be deleted or test entries in the Bizinta system."),
        ("Status Colour Coding",
         "Red = CRITICAL Year 5+ (rotate immediately). "
         "Orange = WARNING Year 4 (plan rotation - 1 year remaining). "
         "Yellow = MONITOR Year 3 (2 years remaining). "
         "Light green = OK Year 2. Dark green = OK Year 1."),
        ("Data Sources",
         "PCAOB: FirmFilings.csv bulk dataset downloaded from pcaobus.org (official PCAOB source, "
         "Firm ID 6651, Latest Form AP Filing = 1 filter applied to deduplicate by fiscal year). "
         "Bizinta: Live GraphQL API call at time of report generation. "
         "EQR: Bizinta custom field Property ID 45173303."),
    ]
    for ri, (title, body) in enumerate(NOTES, 2):
        c1 = ws7.cell(ri, 1, title)
        c1.font = Font(name=F, bold=True, size=10, color=DARK_BLUE)
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c1.fill = PatternFill("solid", fgColor="DEEAF1")
        c1.border = tbdr("CCCCCC")
        c2 = ws7.cell(ri, 2, body)
        c2.font = Font(name=F, size=10)
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c2.border = tbdr("CCCCCC")
        ws7.row_dimensions[ri].height = 65

    # -----------------------------------------------------------------------
    # Sheet 8: Raw PCAOB Filings (all firm ID 6651 records)
    # -----------------------------------------------------------------------
    ws8 = wb.create_sheet("Raw PCAOB Filings")
    ws8.sheet_properties.tabColor = "C0C0C0"
    ws8.merge_cells("A1:J1")
    ws8["A1"] = (f"RAW PCAOB FORM AP FILINGS - Firm ID 6651 (all name variants) | "
                 f"{len(raw_pcaob_rows)} total rows | {run_date_str}")
    ws8["A1"].font = Font(name=F, bold=True, size=11, color=DARK_BLUE)
    ws8["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws8.row_dimensions[1].height = 22
    note_row(ws8, 2, 10,
        "All rows from FirmFilings.csv where Firm ID = 6651. "
        "Includes all firm name variants: Benjamin & Co, Kreit & Chiu CPA LLP, Paris Kreit & Chiu CPA LLP. "
        "This is the unprocessed source data before deduplication and normalisation.",
        DARK_BLUE, "F0F0F0", height=24)
    if raw_pcaob_rows:
        raw_headers = list(raw_pcaob_rows[0].keys())
        ws8.row_dimensions[3].height = 32
        for c, h in enumerate(raw_headers, 1):
            hdr(ws8.cell(3, c, h), bg=GREY)
            ws8.column_dimensions[get_column_letter(c)].width = max(12, min(40, len(h)+4))
        for ri, row in enumerate(raw_pcaob_rows, 4):
            for ci, h in enumerate(raw_headers, 1):
                cell = ws8.cell(ri, ci, row.get(h,""))
                cell.font = Font(name=F, size=9)
                cell.border = tbdr("EEEEEE")
                cell.alignment = Alignment(vertical="center")
            ws8.row_dimensions[ri].height = 14
    ws8.freeze_panes = "A4"

    # -----------------------------------------------------------------------
    # Sheet 9: Raw Bizinta Data (with mapping flag per Section 13)
    # -----------------------------------------------------------------------
    ws9 = wb.create_sheet("Raw Bizinta Data")
    ws9.sheet_properties.tabColor = "FFC000"
    ws9.merge_cells("A1:F1")
    ws9["A1"] = f"RAW BIZINTA CLIENT DATA - Live API pull | {len(bizinta_data)} clients | {run_date_str}"
    ws9["A1"].font = Font(name=F, bold=True, size=11, color=DARK_BLUE)
    ws9["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws9.row_dimensions[1].height = 22
    note_row(ws9, 2, 6,
        "Active clients with no PCAOB mapping are flagged as 'Active in Bizinta - PCAOB mapping not found'. "
        "These should be manually verified to determine whether they are PCAOB-registered clients.",
        "7F3F00", "FFF2CC", height=28)
    BIZ_HDRS = ["Bizinta Display Name", "Status", "EP (Client Manager)",
                "EQR (Custom Field)", "PCAOB Mapped Name", "Mapping / Filing Status"]
    ws9.row_dimensions[3].height = 32
    for c, h in enumerate(BIZ_HDRS, 1): hdr(ws9.cell(3, c, h), bg="7F6000")
    CW9 = [36, 14, 24, 24, 40, 36]
    for i, w in enumerate(CW9, 1): ws9.column_dimensions[get_column_letter(i)].width = w

    for ri, (biz_name, info) in enumerate(sorted(bizinta_data.items()), 4):
        mapped = BIZINTA_TO_PCAOB.get(biz_name)
        has_filing = mapped and mapped in all_pcaob_issuers
        if info.get("status") == "Active" and not mapped:
            flag = "Active in Bizinta - PCAOB mapping not found"
            flag_bg = "FFF2CC"
        elif info.get("status") == "Active" and mapped and not has_filing:
            flag = "Active in Bizinta - mapped but no PCAOB filing found"
            flag_bg = "FCE4D6"
        elif info.get("status") == "Active" and has_filing:
            flag = "Active in Bizinta - PCAOB filing confirmed"
            flag_bg = "E2EFDA"
        else:
            flag = "Inactive / not active"
            flag_bg = "F0F0F0"

        rv = [biz_name, info.get("status",""), info.get("ep",""),
              info.get("eqr",""), mapped if mapped else "(no mapping)", flag]
        for ci, v in enumerate(rv, 1):
            cell = ws9.cell(ri, ci, v)
            cell.font = Font(name=F, size=10)
            cell.border = tbdr("DDDDDD")
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in [1, 5, 6]))
            if ci == 2:
                if v == "Active":
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")
                    cell.font = Font(name=F, size=10, color="375623")
                elif v == "Inactive":
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    cell.font = Font(name=F, size=10, color="7F3F00")
            elif ci == 6:
                cell.fill = PatternFill("solid", fgColor=flag_bg)
                if "not found" in flag or "not active" in flag:
                    cell.font = Font(name=F, size=9, italic=True, color="7F3F00")
                elif "no filing" in flag:
                    cell.font = Font(name=F, size=9, italic=True, color="7F0000")
            elif ci == 5 and not mapped:
                cell.font = Font(name=F, size=10, italic=True, color="888888")
        ws9.row_dimensions[ri].height = 16
    ws9.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# HTML dashboard
# ---------------------------------------------------------------------------
def build_html(enriched, dashboard, pcaob_to_eqr, bizinta_data, run_date_str):
    gf = detect_gap_years(enriched)

    def jsd(o):
        s = json.dumps(o, ensure_ascii=False)
        s = s.replace("&", "\\u0026").replace("<", "\\u003c").replace(">", "\\u003e")
        return s

    dash_js = jsd([{
        "ep":       r["ep"],
        "bizEp":    biz_ep_for_pcaob(r["issuer"], bizinta_data),
        "issuer":   r["issuer"],
        "startYr":  r["start_yr"],
        "yr":       r["year"],
        "consec":   r["consec"],
        "left":     r["yrs_left"],
        "sc":       r["sc"],
        "sl":       r["sl"],
        "signer":   r["signer"],
        "filed":    r["filed"][:10],
        "fye":      r["fye"][:10] if r["fye"] else "",
        "eqr":      pcaob_to_eqr.get(r["issuer"], ""),
        "allYrs":   r["all_yrs"],
        "calcNote": r["calc_note"],
        "hasGap":   any((r["ep"], r["issuer"], yr) in gf for yr in r["all_yrs"]),
    } for r in dashboard])

    all_js = jsd([{
        "ep":       r["ep"],
        "bizEp":    biz_ep_for_pcaob(r["issuer"], bizinta_data),
        "issuer":   r["issuer"],
        "startYr":  r["start_yr"],
        "yr":       r["year"],
        "consec":   r["consec"],
        "left":     r["yrs_left"],
        "sc":       r["sc"],
        "sl":       r["sl"],
        "signer":   r["signer"],
        "filed":    r["filed"][:10],
        "eqr":      pcaob_to_eqr.get(r["issuer"], ""),
        "calcNote": r["calc_note"],
        "hasGap":   (r["ep"], r["issuer"], r["year"]) in gf,
    } for r in enriched])

    eps    = sorted(set(r["ep"] for r in dashboard))
    yrs    = sorted(set(r["year"] for r in enriched))
    ep_opts = "".join(f'<option value="{e}">{e}</option>' for e in eps)
    yr_opts = "".join(f'<option value="{y}">{y}</option>' for y in yrs)

    html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>PCAOB Rotation Dashboard - Kreit &amp; Chiu CPA LLP</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;background:#f0f4f8;color:#1a1a2e;font-size:13px}}
.hdr{{background:linear-gradient(135deg,#1F4E79,#2E75B6);color:#fff;padding:18px 28px;display:flex;justify-content:space-between;align-items:center}}
.hdr h1{{font-size:18px;font-weight:700}}.hdr .sub{{font-size:11px;opacity:.85;margin-top:3px}}
.hdr .badge{{background:rgba(255,255,255,.18);border-radius:8px;padding:8px 16px;text-align:center}}
.hdr .badge .n{{font-size:20px;font-weight:700}}
.stats{{display:flex;gap:12px;padding:14px 24px;background:#fff;border-bottom:1px solid #dde3ea;flex-wrap:wrap}}
.sc{{flex:1;min-width:110px;border-radius:8px;padding:12px 14px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.07)}}
.sc .n{{font-size:26px;font-weight:700}}.sc .l{{font-size:10px;text-transform:uppercase;letter-spacing:.5px;margin-top:3px;opacity:.75}}
.s5c{{background:#fff0f0;color:#cc0000;border-left:4px solid #FF0000}}
.s4c{{background:#fff4e5;color:#cc5500;border-left:4px solid #FF8C00}}
.s3c{{background:#fffbe5;color:#806600;border-left:4px solid #FFD700}}
.okc{{background:#f0fff4;color:#006600;border-left:4px solid #00B050}}
.ttc{{background:#f0f4ff;color:#1F4E79;border-left:4px solid #1F4E79}}
.ctrl{{padding:11px 24px;background:#fff;border-bottom:1px solid #dde3ea;display:flex;gap:12px;align-items:center;flex-wrap:wrap}}
.ctrl label{{font-size:11px;font-weight:700;color:#555}}
.ctrl select,.ctrl input{{border:1px solid #ccc;border-radius:4px;padding:5px 10px;font-size:12px;background:#fff}}
.tabs{{margin-left:auto;display:flex;gap:6px}}
.tb{{padding:6px 14px;border-radius:4px;border:1px solid #ccc;font-size:12px;cursor:pointer;background:#fff;font-weight:500}}
.tb.active{{background:#1F4E79;color:#fff;border-color:#1F4E79}}
.main{{padding:14px 24px}}
.note{{background:#e8f4fd;border-left:4px solid #2E75B6;border-radius:0 6px 6px 0;padding:10px 16px;margin-bottom:8px;font-size:11px;color:#1a3a5c;line-height:1.6}}
.note-warn{{background:#fff4e5;border-left:4px solid #FF8C00;border-radius:0 6px 6px 0;padding:10px 16px;margin-bottom:8px;font-size:11px;color:#7F3F00;line-height:1.6}}
.note-danger{{background:#fff0f0;border-left:4px solid #FF0000;border-radius:0 6px 6px 0;padding:10px 16px;margin-bottom:14px;font-size:11px;color:#7F0000;line-height:1.6}}
.leg{{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:12px;padding:9px 14px;background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
.li{{display:flex;align-items:center;gap:6px;font-size:11px}}
.ld{{width:13px;height:13px;border-radius:3px;flex-shrink:0}}
.stit{{font-size:13px;font-weight:700;color:#1F4E79;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #2E75B6;display:flex;align-items:center;gap:8px}}
.cnt{{background:#2E75B6;color:#fff;font-size:10px;padding:2px 7px;border-radius:10px}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 6px rgba(0,0,0,.08);margin-bottom:18px}}
th{{background:#1F4E79;color:#fff;padding:10px 12px;text-align:left;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;white-space:nowrap}}
td{{padding:9px 12px;border-bottom:1px solid #edf0f5;font-size:12px;vertical-align:middle}}
tr:last-child td{{border-bottom:none}}tr:hover td{{background:#f7f9fc}}
.pill{{display:inline-block;padding:3px 10px;border-radius:12px;font-size:10px;font-weight:700;white-space:nowrap}}
.p5{{background:#FF0000;color:#fff}}.p4{{background:#FF8C00;color:#fff}}
.p3{{background:#FFD700;color:#333}}.p2{{background:#92D050;color:#333}}.p1{{background:#00B050;color:#fff}}
.bar{{display:flex;gap:3px;align-items:center}}
.bk{{width:17px;height:17px;border-radius:3px}}
.f5{{background:#FF0000}}.f4{{background:#FF8C00}}.f3{{background:#FFD700}}
.f2{{background:#92D050}}.f1{{background:#00B050}}.fe{{background:#e0e0e0}}
.eqr-biz{{display:inline-block;background:#E2EFDA;border:1px solid #70AD47;border-radius:4px;padding:2px 7px;font-size:11px;color:#375623;font-weight:600}}
.biz-ep{{display:inline-block;background:#E2EFDA;border:1px solid #70AD47;border-radius:4px;padding:2px 7px;font-size:11px;color:#375623}}
.hidden{{display:none}}.nr{{text-align:center;padding:36px;color:#888;font-style:italic}}
.start-yr{{font-size:13px;font-weight:700;color:#1F4E79;background:#DEEAF1;padding:2px 7px;border-radius:4px;display:inline-block}}
.clickable{{cursor:pointer}}.clickable:hover td{{background:#EBF5FB !important}}
.gap-row td{{background:#FFF0E0 !important}}
.gap-badge{{display:inline-block;background:#FF8C00;color:#fff;font-size:9px;font-weight:700;padding:1px 5px;border-radius:3px;margin-left:4px}}
.modal-overlay{{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center}}
.modal-overlay.open{{display:flex}}
.modal{{background:#fff;border-radius:10px;padding:28px;max-width:700px;width:90%;max-height:82vh;overflow-y:auto;box-shadow:0 8px 32px rgba(0,0,0,.25)}}
.modal-hdr{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:18px;padding-bottom:14px;border-bottom:2px solid #1F4E79}}
.modal-hdr h2{{font-size:15px;color:#1F4E79;font-weight:700}}
.modal-close{{background:none;border:none;font-size:20px;cursor:pointer;color:#888;line-height:1;padding:2px 6px}}
.modal-section{{margin-bottom:16px}}
.modal-section h3{{font-size:12px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}}
.modal-grid{{display:grid;grid-template-columns:160px 1fr;gap:6px 12px;font-size:12px}}
.modal-label{{color:#888;font-weight:600}}.modal-value{{color:#1a1a2e}}
.calc-box{{background:#f8f9fa;border-left:3px solid #2E75B6;padding:10px 14px;border-radius:0 6px 6px 0;font-size:11px;color:#1a3a5c;line-height:1.7;margin-top:8px}}
.hist-table{{width:100%;border-collapse:collapse;font-size:11px;margin-top:8px}}
.hist-table th{{background:#f0f4f8;color:#555;padding:6px 10px;text-align:left;font-weight:700;border-bottom:2px solid #dde3ea}}
.hist-table td{{padding:5px 10px;border-bottom:1px solid #f0f0f0;vertical-align:middle}}
.gap-warn{{background:#fff4e5;border:1px solid #FF8C00;border-radius:6px;padding:10px 14px;font-size:11px;color:#7F3F00;margin-top:10px}}
</style></head><body>
<div class="hdr">
  <div><h1>PCAOB Partner Rotation Dashboard</h1>
  <div class="sub">Kreit &amp; Chiu CPA LLP &nbsp;|&nbsp; Firm ID: 6651 &nbsp;|&nbsp; PCAOB AS 1201 - 5-Year Rule &nbsp;|&nbsp; Generated: {run_date_str}</div></div>
  <div class="badge"><div class="n" id="hdr-date"></div><div style="font-size:10px;opacity:.8">Run Date</div></div>
</div>
<div class="stats" id="stats"></div>
<div class="ctrl">
  <label>Partner (Form AP):</label>
  <select id="ep-f" onchange="filter()"><option value="">All Partners</option>{ep_opts}</select>
  <label>Status:</label>
  <select id="st-f" onchange="filter()">
    <option value="">All</option><option value="5">Critical (Yr 5+)</option>
    <option value="4">Warning (Yr 4)</option><option value="3">Monitor (Yr 3)</option>
    <option value="1">OK (Yr 1-2)</option>
  </select>
  <label>Year:</label>
  <select id="yr-f" onchange="filter()"><option value="">All Years</option>{yr_opts}</select>
  <label>Search:</label>
  <input id="srch" type="text" placeholder="Issuer or partner..." oninput="filter()" style="width:180px">
  <label style="margin-left:8px"><input type="checkbox" id="gap-only" onchange="filter()"> Gap flags only</label>
  <div class="tabs">
    <button class="tb active" onclick="tab('dash')">Rotation Dashboard</button>
    <button class="tb" onclick="tab('hist')">Full Filing History</button>
  </div>
</div>
<div class="main">
<div class="note">
  <strong>SCOPE:</strong> This dashboard shows only clients that are <strong>Active in Bizinta AND have a PCAOB Form AP filing</strong> (Firm ID 6651).
  Clients active in Bizinta with no PCAOB filing are excluded here and listed separately for manual review.
  Clients with filings but no longer active in Bizinta are excluded as the rotation chain is considered broken.
  &nbsp;<strong>EQR Assumption B (CRITICAL):</strong> EQR tenure assumed to start same year as EP. Human review required where internal records differ.
</div>
<div class="note-danger">
  <strong>Gap-Year Assumption A (CRITICAL):</strong> A year with no Form AP filing resets the consecutive count to Year 1 on return.
  Rows marked <span class="gap-badge">GAP</span> indicate a gap-then-return pattern requiring human review.
  An engineered gap to reset the clock constitutes circumvention under PCAOB rules.
</div>
<div id="t-dash">
  <div class="leg">
    <span style="font-size:11px;font-weight:700;color:#555;margin-right:4px">Legend:</span>
    <div class="li"><div class="ld" style="background:#FF0000"></div>Year 5+ CRITICAL</div>
    <div class="li"><div class="ld" style="background:#FF8C00"></div>Year 4 WARNING</div>
    <div class="li"><div class="ld" style="background:#FFD700"></div>Year 3 MONITOR</div>
    <div class="li"><div class="ld" style="background:#92D050"></div>Year 2 OK</div>
    <div class="li"><div class="ld" style="background:#00B050"></div>Year 1 OK</div>
    <div class="li" style="margin-left:12px"><span style="font-size:10px;color:#777">Click any row for drill-down</span></div>
  </div>
  <div class="stit">Current Engagement Status (Active Bizinta Clients Only) <span class="cnt" id="dc">0</span></div>
  <table><thead><tr>
    <th>EP (Form AP)</th><th>EP (Bizinta)</th><th>Issuer / Client</th>
    <th style="text-align:center">EP Start Yr</th>
    <th style="text-align:center">Audit Yr</th><th style="text-align:center">Consec Yrs</th>
    <th style="text-align:center">Yrs Left</th><th>Rotation Status</th>
    <th>Signer</th><th>EQR</th>
  </tr></thead><tbody id="db"></tbody></table>
</div>
<div id="t-hist" class="hidden">
  <div class="stit">All Form AP Filings - Firm ID 6651 <span class="cnt" id="hc">0</span></div>
  <table><thead><tr>
    <th>Filed</th><th style="text-align:center">Yr</th>
    <th>EP (Form AP)</th><th>EP (Bizinta)</th><th>Issuer / Client</th>
    <th style="text-align:center">Rotation Yr</th><th>Status</th>
    <th>Signer</th><th>EQR</th>
  </tr></thead><tbody id="hb"></tbody></table>
</div>
</div>

<div class="modal-overlay" id="modal" onclick="closeModal(event)">
  <div class="modal">
    <div class="modal-hdr">
      <h2 id="modal-title">Engagement Detail</h2>
      <button class="modal-close" onclick="closeModalBtn()">x</button>
    </div>
    <div id="modal-body"></div>
  </div>
</div>

<script>
const D={dash_js};const A={all_js};
function pc(sc){{return sc>=5?'p5':sc==4?'p4':sc==3?'p3':sc==2?'p2':'p1'}}
function bar(c){{let h='<div class="bar">';for(let i=1;i<=5;i++)h+=`<div class="bk ${{i<=c?'f'+Math.min(c,5):'fe'}}"></div>`;return h+'</div>';}}
function slabel(sc){{const m={{5:'CRITICAL Yr 5+',4:'WARNING Yr 4',3:'MONITOR Yr 3',2:'OK Yr 2',1:'OK Yr 1'}};return m[Math.min(sc,5)];}}
function esc(s){{const d=document.createElement('div');d.textContent=String(s||'');return d.innerHTML;}}

function renderStats(){{
  const c5=D.filter(r=>r.sc>=5).length,c4=D.filter(r=>r.sc==4).length,c3=D.filter(r=>r.sc==3).length,ok=D.filter(r=>r.sc<=2).length;
  document.getElementById('stats').innerHTML=
    `<div class="sc s5c"><div class="n">${{c5}}</div><div class="l">Critical (Yr 5+)</div></div>`+
    `<div class="sc s4c"><div class="n">${{c4}}</div><div class="l">Warning (Yr 4)</div></div>`+
    `<div class="sc s3c"><div class="n">${{c3}}</div><div class="l">Monitor (Yr 3)</div></div>`+
    `<div class="sc okc"><div class="n">${{ok}}</div><div class="l">OK (Yr 1-2)</div></div>`+
    `<div class="sc ttc"><div class="n">${{D.length}}</div><div class="l">Active Engagements</div></div>`;
}}

function openModal(r){{
  const hist=A.filter(x=>x.ep===r.ep&&x.issuer===r.issuer).sort((a,b)=>a.yr-b.yr);
  const histRows=hist.map(x=>`<tr><td>${{x.yr}}</td><td>${{x.filed}}</td><td>${{bar(x.consec)}} Yr ${{x.consec}}</td><td><span class="pill ${{pc(x.sc)}}">${{slabel(x.sc)}}</span></td><td>${{x.hasGap?'<span class="gap-badge">GAP</span>':''}}</td></tr>`).join('');
  const gapWarn=r.hasGap?`<div class="gap-warn"><strong>GAP FLAG:</strong> A gap year was detected. Count restarted after gap. Human review required.</div>`:'';
  const leftHtml=r.left===0?'<strong style="color:#cc0000">ROTATE NOW</strong>':`<strong style="color:${{r.left===1?'#cc5500':'#006600'}}">${{r.left}} year(s)</strong>`;
  document.getElementById('modal-title').textContent=r.ep+' \u2013 '+r.issuer;
  document.getElementById('modal-body').innerHTML=`
    <div class="modal-section"><h3>Engagement Summary</h3>
    <div class="modal-grid">
      <span class="modal-label">EP (Form AP)</span><span class="modal-value"><strong>${{esc(r.ep)}}</strong></span>
      <span class="modal-label">EP (Bizinta)</span><span class="modal-value">${{r.bizEp?`<span class="biz-ep">${{esc(r.bizEp)}}</span>`:'<span style="color:#aaa">Not in Bizinta</span>'}}</span>
      <span class="modal-label">Issuer</span><span class="modal-value">${{esc(r.issuer)}}</span>
      <span class="modal-label">EP Start Year</span><span class="modal-value"><span class="start-yr">${{r.startYr||r.yr}}</span></span>
      <span class="modal-label">Latest Audit Year</span><span class="modal-value">${{r.yr}}</span>
      <span class="modal-label">EQR</span><span class="modal-value">${{r.eqr?`<span class="eqr-biz">${{esc(r.eqr)}}</span>`:'Not in Bizinta'}}</span>
      <span class="modal-label">Signer</span><span class="modal-value">${{esc(r.signer||'-')}}</span>
      <span class="modal-label">Status</span><span class="modal-value"><span class="pill ${{pc(r.sc)}}">${{slabel(r.sc)}}</span></span>
      <span class="modal-label">Years Remaining</span><span class="modal-value">${{leftHtml}}</span>
    </div></div>
    <div class="modal-section"><h3>Calculation Logic</h3>
    <div class="calc-box">${{esc(r.calcNote)}}</div>${{gapWarn}}</div>
    <div class="modal-section"><h3>Full Filing History (${{hist.length}} year(s))</h3>
    <table class="hist-table"><thead><tr><th>Audit Yr</th><th>Filed</th><th>Consecutive</th><th>Status</th><th></th></tr></thead>
    <tbody>${{histRows}}</tbody></table></div>`;
  document.getElementById('modal').classList.add('open');
}}
function closeModal(e){{if(e.target===document.getElementById('modal'))closeModalBtn();}}
function closeModalBtn(){{document.getElementById('modal').classList.remove('open');}}
document.addEventListener('keydown',e=>{{if(e.key==='Escape')closeModalBtn();}});

function renderDash(data){{
  const tb=document.getElementById('db');document.getElementById('dc').textContent=data.length;
  if(!data.length){{tb.innerHTML='<tr><td colspan="10" class="nr">No records match.</td></tr>';return;}}
  tb.innerHTML=data.map((r,i)=>`<tr data-idx="${{i}}" class="clickable${{r.hasGap?' gap-row':''}}">
    <td><strong>${{esc(r.ep)}}</strong></td>
    <td>${{r.bizEp?`<span class="biz-ep">${{esc(r.bizEp)}}</span>`:'<span style="color:#aaa;font-size:11px">-</span>'}}</td>
    <td>${{esc(r.issuer)}}${{r.hasGap?'<span class="gap-badge">GAP</span>':''}}</td>
    <td style="text-align:center"><span class="start-yr">${{r.startYr||r.yr}}</span></td>
    <td style="text-align:center">${{r.yr}}</td>
    <td style="text-align:center">${{bar(r.consec)}} <span style="font-size:10px;color:#666">Yr ${{r.consec}}</span></td>
    <td style="text-align:center"><strong style="color:${{r.left==0?'#cc0000':r.left==1?'#cc5500':'#006600'}}">${{r.left==0?'ROTATE NOW':r.left}}</strong></td>
    <td><span class="pill ${{pc(r.sc)}}">${{slabel(r.sc)}}</span></td>
    <td style="color:#555;font-size:11px">${{esc(r.signer)}}</td>
    <td>${{r.eqr?`<span class="eqr-biz">${{esc(r.eqr)}}</span>`:'<span style="color:#aaa;font-size:11px">-</span>'}}</td>
  </tr>`).join('');
  const lm={{}};data.forEach((r,i)=>lm[i]=r);
  document.getElementById('db').querySelectorAll('tr[data-idx]').forEach(tr=>{{
    tr.addEventListener('click',()=>openModal(lm[+tr.dataset.idx]));
  }});
}}

function renderHist(data){{
  const tb=document.getElementById('hb');document.getElementById('hc').textContent=data.length;
  if(!data.length){{tb.innerHTML='<tr><td colspan="9" class="nr">No records match.</td></tr>';return;}}
  tb.innerHTML=data.map((r,i)=>`<tr data-idx="${{i}}" class="clickable${{r.hasGap?' gap-row':''}}">
    <td>${{r.filed}}</td><td style="text-align:center">${{r.yr}}</td>
    <td><strong>${{esc(r.ep)}}</strong></td>
    <td>${{r.bizEp?`<span class="biz-ep">${{esc(r.bizEp)}}</span>`:'<span style="color:#aaa;font-size:10px">-</span>'}}</td>
    <td>${{esc(r.issuer)}}${{r.hasGap?'<span class="gap-badge">GAP</span>':''}}</td>
    <td style="text-align:center">${{bar(r.consec)}}</td>
    <td><span class="pill ${{pc(r.sc)}}">${{slabel(r.sc)}}</span></td>
    <td style="color:#555;font-size:11px">${{esc(r.signer)}}</td>
    <td>${{r.eqr?`<span class="eqr-biz">${{esc(r.eqr)}}</span>`:''}}</td>
  </tr>`).join('');
  const lm={{}};data.forEach((r,i)=>lm[i]=r);
  document.getElementById('hb').querySelectorAll('tr[data-idx]').forEach(tr=>{{
    tr.addEventListener('click',()=>openModal(lm[+tr.dataset.idx]));
  }});
}}

function filter(){{
  const ep=document.getElementById('ep-f').value,
        st=document.getElementById('st-f').value,
        yr=document.getElementById('yr-f').value,
        q=document.getElementById('srch').value.toLowerCase(),
        gapOnly=document.getElementById('gap-only').checked;
  const fd=r=>{{
    if(ep&&r.ep!==ep)return false;
    if(st){{const c=parseInt(st);if(c==1&&r.sc>2)return false;if(c>1&&r.sc!==c)return false;}}
    if(yr&&String(r.yr)!==yr)return false;
    if(q&&!r.ep.toLowerCase().includes(q)&&!r.issuer.toLowerCase().includes(q))return false;
    if(gapOnly&&!r.hasGap)return false;
    return true;
  }};
  renderDash(D.filter(fd));renderHist(A.filter(fd));
}}

function tab(t){{
  document.getElementById('t-dash').classList.toggle('hidden',t!=='dash');
  document.getElementById('t-hist').classList.toggle('hidden',t!=='hist');
  document.querySelectorAll('.tb').forEach((b,i)=>b.classList.toggle('active',(i==0&&t=='dash')||(i==1&&t=='hist')));
}}

document.getElementById('hdr-date').textContent=new Date().toLocaleDateString('en-GB',{{day:'2-digit',month:'short',year:'numeric'}});
renderStats();renderDash(D);renderHist(A);
</script></body></html>"""
    return html


# ---------------------------------------------------------------------------
# Email alert
# ---------------------------------------------------------------------------
def build_email_alert(dashboard, run_date_str):
    critical = [r for r in dashboard if r["sc"] >= 5]
    warning  = [r for r in dashboard if r["sc"] == 4]
    monitor  = [r for r in dashboard if r["sc"] == 3]
    subject  = (f"PCAOB Rotation Alert - {run_date_str} | "
                f"{len(critical)} Critical, {len(warning)} Warning, {len(monitor)} Monitor")

    def rows_html(items, bg, fg):
        if not items:
            return '<tr><td colspan="5" style="color:#888;font-style:italic;padding:10px 12px">None</td></tr>'
        return "".join(f"""<tr>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;font-weight:600">{r['ep']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #eee">{r['issuer']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center">{r['year']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center">
            <span style="background:{bg};color:{fg};padding:3px 10px;border-radius:12px;font-size:11px;font-weight:700">Year {r['consec']}</span>
          </td>
          <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center">
            {'<strong style="color:#cc0000">ROTATE NOW</strong>' if r['yrs_left']==0 else f'{r["yrs_left"]} yr(s)'}
          </td></tr>""" for r in items)

    th = 'style="background:#1F4E79;color:#fff;padding:9px 12px;text-align:left;font-size:11px;font-weight:700;text-transform:uppercase"'
    html_body = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:0">
<div style="max-width:700px;margin:24px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1)">
  <div style="background:linear-gradient(135deg,#1F4E79,#2E75B6);padding:24px 28px;color:#fff">
    <h1 style="margin:0;font-size:20px">PCAOB Partner Rotation Alert</h1>
    <p style="margin:6px 0 0;font-size:12px;opacity:.85">Kreit &amp; Chiu CPA LLP &nbsp;|&nbsp; Firm ID 6651 &nbsp;|&nbsp; Active Bizinta clients scope &nbsp;|&nbsp; {run_date_str}</p>
  </div>
  <div style="display:flex;border-bottom:3px solid #eee">
    <div style="flex:1;text-align:center;padding:16px;border-right:1px solid #eee"><div style="font-size:28px;font-weight:700;color:#cc0000">{len(critical)}</div><div style="font-size:11px;color:#666;margin-top:3px;text-transform:uppercase">Critical (Yr 5+)</div></div>
    <div style="flex:1;text-align:center;padding:16px;border-right:1px solid #eee"><div style="font-size:28px;font-weight:700;color:#cc5500">{len(warning)}</div><div style="font-size:11px;color:#666;margin-top:3px;text-transform:uppercase">Warning (Yr 4)</div></div>
    <div style="flex:1;text-align:center;padding:16px;border-right:1px solid #eee"><div style="font-size:28px;font-weight:700;color:#806600">{len(monitor)}</div><div style="font-size:11px;color:#666;margin-top:3px;text-transform:uppercase">Monitor (Yr 3)</div></div>
    <div style="flex:1;text-align:center;padding:16px"><div style="font-size:28px;font-weight:700;color:#1F4E79">{len(dashboard)}</div><div style="font-size:11px;color:#666;margin-top:3px;text-transform:uppercase">Active Engagements</div></div>
  </div>
  <div style="padding:24px 28px">
    <h2 style="font-size:14px;color:#cc0000;margin:0 0 10px;padding-bottom:6px;border-bottom:2px solid #FF0000">CRITICAL - Rotate Immediately (Year 5+)</h2>
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;font-size:12px">
      <tr><th {th}>Partner</th><th {th}>Client</th><th {th} style="text-align:center">Yr</th><th {th} style="text-align:center">Status</th><th {th} style="text-align:center">Left</th></tr>
      {rows_html(critical,'#FF0000','#fff')}
    </table>
    <h2 style="font-size:14px;color:#cc5500;margin:0 0 10px;padding-bottom:6px;border-bottom:2px solid #FF8C00">WARNING - Plan Rotation (Year 4)</h2>
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;font-size:12px">
      <tr><th {th}>Partner</th><th {th}>Client</th><th {th} style="text-align:center">Yr</th><th {th} style="text-align:center">Status</th><th {th} style="text-align:center">Left</th></tr>
      {rows_html(warning,'#FF8C00','#fff')}
    </table>
    <h2 style="font-size:14px;color:#806600;margin:0 0 10px;padding-bottom:6px;border-bottom:2px solid #FFD700">MONITOR - On Watch (Year 3)</h2>
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;font-size:12px">
      <tr><th {th}>Partner</th><th {th}>Client</th><th {th} style="text-align:center">Yr</th><th {th} style="text-align:center">Status</th><th {th} style="text-align:center">Left</th></tr>
      {rows_html(monitor,'#FFD700','#333')}
    </table>
    <div style="background:#e8f4fd;border-left:4px solid #2E75B6;border-radius:0 6px 6px 0;padding:12px 16px;font-size:11px;color:#1a3a5c;line-height:1.7">
      Dashboard scope: Active Bizinta clients with PCAOB Form AP filings under Firm ID 6651 only.<br>
      EQR tenure assumed same start year as EP. Human review required where internal records differ.<br>
      Cooling-off period tracking is a separate enhancement not yet implemented.
    </div>
  </div>
  <div style="background:#f4f6f9;padding:12px 28px;font-size:10px;color:#999;text-align:center">
    Auto-generated by PCAOB Rotation Build Service &nbsp;|&nbsp; Kreit &amp; Chiu CPA LLP &nbsp;|&nbsp; Confidential - do not forward externally.
  </div>
</div></body></html>"""

    return subject, html_body, {
        "critical_count": len(critical), "warning_count": len(warning),
        "monitor_count": len(monitor), "total": len(dashboard),
        "critical": [{"ep":r["ep"],"issuer":r["issuer"],"year":r["year"],"consec":r["consec"]} for r in critical],
        "warning":  [{"ep":r["ep"],"issuer":r["issuer"],"year":r["year"],"consec":r["consec"]} for r in warning],
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/health")
def health():
    return jsonify({"status": "ok", "service": "pcaob-build-service-v3"})

@app.route("/build", methods=["POST"])
@require_api_key
def build():
    run_date     = date.today()
    run_date_str = run_date.strftime("%d %b %Y")
    fname_date   = run_date.strftime("%Y-%m-%d")

    try:
        bizinta_data = fetch_bizinta()
    except Exception as e:
        return jsonify({"error": f"Bizinta API failed: {e}"}), 502

    try:
        records, raw_pcaob_rows = fetch_pcaob()
    except Exception as e:
        return jsonify({"error": f"PCAOB filter service failed: {e}"}), 502

    # EQR lookup: pcaob issuer name -> eqr string
    pcaob_to_eqr = {}
    for biz_name, biz in bizinta_data.items():
        pcaob_name = BIZINTA_TO_PCAOB.get(biz_name)
        if pcaob_name and biz.get("eqr"):
            pcaob_to_eqr[pcaob_name] = biz["eqr"]

    enriched, dashboard = build_rotation(records, bizinta_data)

    excel_bytes  = build_excel(enriched, dashboard, pcaob_to_eqr, bizinta_data, raw_pcaob_rows, run_date_str)
    html_str     = build_html(enriched, dashboard, pcaob_to_eqr, bizinta_data, run_date_str)
    email_subj, email_html, summary = build_email_alert(dashboard, run_date_str)

    return jsonify({
        "excel_b64":      base64.b64encode(excel_bytes).decode("utf-8"),
        "html_b64":       base64.b64encode(html_str.encode("utf-8")).decode("utf-8"),
        "excel_filename": f"PCAOB_Rotation_Tracker_KreitChiu_{fname_date}.xlsx",
        "html_filename":  f"PCAOB_Rotation_Dashboard_KreitChiu_{fname_date}.html",
        "email_subject":  email_subj,
        "email_html":     email_html,
        "summary":        summary,
        "run_date":       fname_date,
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
